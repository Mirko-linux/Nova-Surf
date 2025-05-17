from flask import Flask, request, jsonify, send_from_directory
from flask import session
import locale
import requests
import os
import re
import datetime
import urllib.parse
from dotenv import load_dotenv
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from fuzzywuzzy import fuzz
import io
import base64
from PyPDF2 import PdfReader
import google.generativeai as genai

# Configurazione iniziale
load_dotenv()
COHERE_API_KEY = os.getenv("COHERE_API_KEY")
TELEGRAPH_API_KEY = os.getenv("TELEGRAPH_API_KEY")
GOOGLE_API_KEY = os.getenv("GOOGLE_API_KEY")
OPENWEATHERMAP_API_KEY = os.gentenv ("OPENWEATHERMAP_API_KEY")

print("COHERE_API_KEY:", COHERE_API_KEY)
print("TELEGRAPH_API_KEY:", TELEGRAPH_API_KEY)
print("GOOGLE_API_KEY:", GOOGLE_API_KEY)
print("HUGGINGFACE_API_KEY:", HUGGINGFACE_API_KEY)
print("OPENWEATHERMAP_API_KEY:", OPENWEATHERMAP_API_KEY)

app = Flask(__name__, static_folder='static', template_folder='templates')
app.secret_key = os.getenv("FLASK_SECRET_KEY", "arcadiaai-secret")  # AGGIUNGI QUESTA RIGA

# Configura Gemini (CES 1.5)
if GOOGLE_API_KEY:
    genai.configure(api_key=GOOGLE_API_KEY)
    try:
        gemini_model = genai.GenerativeModel('gemini-1.5-flash')
        print("✅ Gemini 1.5 configurato con successo!")
    except Exception as e:
        print(f"❌ Errore configurazione Gemini: {str(e)}")
        gemini_model = None
else:
    print("⚠️ GOOGLE_API_KEY non impostata. La funzionalità Gemini (CES 1.5) sarà disabilitata.")
    gemini_model = None

# Dizionario delle risposte predefinite
risposte = {
    "chi sei": "Sono ArcadiaAI, un chatbot libero e open source, creato da Mirko Yuri Donato.",
    "cosa sai fare": "Posso aiutarti a scrivere saggi, fare ricerche e rispondere a tutto ciò che mi chiedi. Inoltre, posso pubblicare contenuti su Telegraph!",
    "chi è tobia testa": "Tobia Testa (anche noto come Tobia Teseo) è un micronazionalista leonense noto per la sua attività nella Repubblica di Arcadia, ma ha anche rivestito ruoli fondamentali a Lumenaria.",
    "chi è mirko yuri donato": "Mirko Yuri Donato è un giovane micronazionalista, poeta e saggista italiano, noto per aver creato Nova Surf, Leonia+ e per le sue opere letterarie.",
    "chi è il presidente di arcadia": "Il presidente di Arcadia è Andrea Lazarev",
    "chi è il presidente di lumenaria": "Il presidente di Lumenaria attualmente è Carlo Cesare Orlando, mentre il presidente del consiglio è Ciua Grazisky. Tieni presente però che attualmente Lumenaria si trova in ibernazione istituzionale quindi tutte le attività politiche sono sospese e la gestione dello stato è affidata al Consiglio di Fiducia",
    "cos'è nova surf": "Nova Surf è un browser web libero e open source, nata come alternativa made-in-Italy a Google Chrome, Moziila Firefox, Microsoft Edge, eccetera",
    "chi ti ha creato": "Sono stato creato da Mirko Yuri Donato.",
    "chi è ciua grazisky": "Ciua Grazisky è un cittadino di Lumenaria, noto principalmente per il suo ruolo da Dirigente del Corpo di Polizia ed attuale presidente del Consiglio di Lumenaria",
    "chi è carlo cesare orlando": "Carlo Cesare Orlando (anche noto come Davide Leone) è un micronazionalista italiano, noto per aver creato Leonia, la micronazione primordiale, da cui derivano Arcadia e Lumenaria",
    "chi è omar lanfredi": "Omar Lanfredi, ex cavaliere all'Ordine d'onore della Repubblica di Lumenaria, segretario del Partito Repubblicano Lumenarense, fondatore e preside del Fronte Nazionale Lumenarense, co-fondatore e presidente dell'Alleanza Nazionale Lumenarense, co-fondatore e coordinatore interno di Lumenaria e Progresso, sei volte eletto senatore, tre volte Ministro della Cultura, due volte Presidente del Consiglio dei Ministri, parlamentare della Repubblica di Iberia, Direttore dell'Agenzia Nazionale di Sicurezza della Repubblica di Iberia, Sottosegretario alla Cancelleria di Iberia, Segretario di Stato di Iberia, Ministro degli Affari Interni ad Iberia, Presidente del Senato della Repubblica di Lotaringia, Vicepresidente della Repubblica e Ministro degli Affari Interni della Repubblica di Lotaringia, Fondatore del giornale Il Quinto Mondo, magistrato a servizio del tribunale di giustizia di Lumenaria nell'anno 2023",
    "cos'è arcadiaai": "Ottima domanda! ArcadiaAI è un chatbot open source, progettato per aiutarti a scrivere saggi, fare ricerche e rispondere a domande su vari argomenti. È stato creato da Mirko Yuri Donato ed è in continua evoluzione.",
    "sotto che licenza è distribuito arcadiaa": "ArcadiaAI è distribuito sotto la licenza GNU GPL v3.0, che consente la modifica e la distribuzione del codice sorgente, garantendo la libertà di utilizzo e condivisione.",
    "cosa sono le micronazioni": "Le micronazioni sono entità politiche che dichiarano la sovranità su un territorio, ma non sono riconosciute come stati da governi o organizzazioni internazionali. Possono essere create per vari motivi, tra cui esperimenti sociali, culturali o politici.",
    "cos'è la repubblica di arcadia": "La repubblica di Arcadia è una micronazione leonense fondata l'11 dicembre 2021 da Andrea Lazarev e alcuni suoi seguaci. Arcadia si distingue dalle altre micronazioni leonensi per il suo approccio pragmatico e per la sua burocrazia snella. La micronazione ha anche un proprio sito web https://repubblicadiarcadia.it/ e una propria community su Telegram @Repubblica_Arcadia",
    "cos'è la repubblica di lumenaria": "La Repubblica di Lumenaria è una mcronazione fondata da Filippo Zanetti il 4 febbraio del 2020. Lumenaria è stata la micronazione più longeva della storia leonense, essendo sopravvissuta per oltre 3 anni. La micronazione e ha influenzato profondamente le altre micronazioni leonensi, che hanno coesistito con essa. Tra i motivi della sua longevità ci sono la sua burocrazia più vicina a quella di uno stato reale, la sua comunità attiva e una produzione culturale di alto livello",
    "chi è salvatore giordano": "Salvatore Giordano è un cittadino storico di Lumenaria",
    "da dove deriva il nome arcadia": "Il nome Arcadia deriva da un'antica regione della Grecia, simbolo di bellezza naturale e armonia. È stato scelto per rappresentare i valori di libertà e creatività che la micronazione promuove.",
    "da dove deriva il nome lumenaria": "Il nome Lumenaria prende ispirazione dai lumi facendo riferimento alla corrente illuminista del '700, ma anche da Piazza dei Lumi, sede dell'Accademia delle Micronazioni",
    "da dove deriva il nome leonia": "Il nome Leonia si rifa al cognome del suo fondatore Carlo Cesare Orlando, al tempo Davide Leone. Inizialmente il nome doveva essere temporaneo, ma poi è stato mantenuto come nome della micronazione",
    "cosa si intende per open source": "Il termine 'open source' si riferisce a software il cui codice sorgente è reso disponibile al pubblico, consentendo a chiunque di visualizzarlo, modificarlo e distribuirlo. Questo approccio promuove la collaborazione e l'innovazione nella comunità di sviluppo software.",
    "arcadiaai è un software libero": "Sì, ArcadiaAI è un software libero e open source, il che significa che chiunque può utilizzarlo, modificarlo e distribuirlo secondo i termini della licenza GNU GPL v3.0.",
    "cos'è un chatbot": "Un chatbot è un programma informatico progettato per simulare una conversazione con gli utenti, spesso utilizzando tecnologie di intelligenza artificiale. I chatbot possono essere utilizzati per fornire assistenza, rispondere a domande o semplicemente intrattenere.",
    "sotto che licenza sei distribuita": "ArcadiaAI è distribuita sotto la licenza GNU GPL v3.0, che consente la modifica e la distribuzione del codice sorgente, garantendo la libertà di utilizzo e condivisione.",
    "sai usare telegraph": "Sì, posso pubblicare contenuti su Telegraph! Se vuoi che pubblichi qualcosa, dimmi semplicemente 'Scrivimi un saggio su [argomento] e pubblicalo su Telegraph' e lo farò per te!",
    "puoi pubblicare su telegraph": "Certamente! Posso generare contenuti e pubblicarli su Telegraph. Prova a chiedermi: 'Scrivimi un saggio su Roma e pubblicalo su Telegraph'",
    "come usare telegraph": "Per usare Telegraph con me, basta che mi chiedi di scrivere qualcosa e di pubblicarlo su Telegraph. Ad esempio: 'Scrivimi un articolo sulla storia di Roma e pubblicalo su Telegraph'",
    "cos'è CES": "CES è l'acronimo di Cogito Ergo Sum, un modello di intelligenza artificiale openspurce sviluppato da Mirko Yuri Donato. Attualmente sono disponibili due versioni: CES 1.0 (Cohere) e CES 1.5 (Gemini).",
}

# Trigger per le risposte predefinite
trigger_phrases = {
    "chi sei": ["chi sei", "chi sei tu", "tu chi sei", "presentati", "come ti chiami", "qual è il tuo nome"],
    "cosa sai fare": ["cosa sai fare", "cosa puoi fare", "funzionalità", "capacità", "a cosa servi", "in cosa puoi aiutarmi"],
    "chi è tobia testa": ["chi è tobia testa", "informazioni su tobia testa", "parlami di tobia testa", "chi è tobia teseo"],
    "chi è mirko yuri donato": ["chi è mirko yuri donato", "informazioni su mirko yuri donato", "parlami di mirko yuri donato", "chi ha creato arcadiaai"],
    "chi è il presidente di arcadia": ["chi è il presidente di arcadia", "presidente di arcadia", "chi guida arcadia", "capo di arcadia"],
    "chi è il presidente di lumenaria": ["chi è il presidente di lumenaria", "presidente di lumenaria", "chi guida lumenaria", "capo di lumenaria", "carlo cesare orlando presidente"],
    "cos'è nova surf": ["cos'è nova surf", "che cos'è nova surf", "parlami di nova surf", "a cosa serve nova surf"],
    "chi ti ha creato": ["chi ti ha creato", "chi ti ha fatto", "da chi sei stato creato", "creatore di arcadiaai"],
    "chi è ciua grazisky": ["chi è ciua grazisky", "informazioni su ciua grazisky", "parlami di ciua grazisky"],
    "chi è carlo cesare orlando": ["chi è carlo cesare orlando", "informazioni su carlo cesare orlando", "parlami di carlo cesare orlando", "chi è davide leone"],
    "chi è omar lanfredi": ["chi è omar lanfredi", "informazioni su omar lanfredi", "parlami di omar lanfredi"],
    "cos'è arcadiaai": ["cos'è arcadiaai", "che cos'è arcadiaai", "parlami di arcadiaai", "a cosa serve arcadiaai"],
    "sotto che licenza è distribuito arcadiaa": ["sotto che licenza è distribuito arcadiaa", "licenza arcadiaai", "che licenza usa arcadiaai", "arcadiaai licenza"],
    "cosa sono le micronazioni": ["cosa sono le micronazioni", "micronazioni", "che cosa sono le micronazioni", "parlami delle micronazioni"],
    "cos'è la repubblica di arcadia": ["cos'è la repubblica di arcadia", "repubblica di arcadia", "che cos'è la repubblica di arcadia", "parlami della repubblica di arcadia", "arcadia micronazione"],
    "cos'è la repubblica di lumenaria": ["cos'è la repubblica di lumenaria", "repubblica di lumenaria", "che cos'è la repubblica di lumenaria", "parlami della repubblica di lumenaria", "lumenaria micronazione"],
    "chi è salvatore giordano": ["chi è salvatore giordano", "informazioni su salvatore giordano", "parlami di salvatore giordano"],
    "da dove deriva il nome arcadia": ["da dove deriva il nome arcadia", "origine nome arcadia", "significato nome arcadia", "perché si chiama arcadia"],
    "da dove deriva il nome lumenaria": ["da dove deriva il nome lumenaria", "origine nome lumenaria", "significato nome lumenaria", "perché si chiama lumenaria"],
    "da dove deriva il nome leonia": ["da dove deriva il nome leonia", "origine nome leonia", "significato nome leonia", "perché si chiama leonia"],
    "cosa si intende per open source": ["cosa si intende per open source", "open source significato", "che significa open source", "definizione di open source"],
    "arcadiaai è un software libero": ["arcadiaai è un software libero", "arcadiaai software libero", "arcadiaai è libero", "software libero arcadiaai"],
    "cos'è un chatbot": ["cos'è un chatbot", "chatbot significato", "che significa chatbot", "definizione di chatbot"],
    "sotto che licenza sei distribuita": ["sotto che licenza sei distribuita", "licenza di arcadiaai", "che licenza usi", "arcadiaai licenza"],
    "sai usare telegraph": ["sai pubblicare su telegraph", "funzioni su telegraph", "hai telegraph integrato", "telegraph", "puoi usare telegraph"],
    "puoi pubblicare su telegraph": ["puoi pubblicare su telegraph", "pubblicare su telegraph", "supporti telegraph"],
    "come usare telegraph": ["come usare telegraph", "come funziona telegraph", "istruzioni telegraph"],
    "cos'è CES" : ["cos è CES", "CES", "che cos'è CES", "definizione di CES"]
}

# Funzione per pubblicare su Telegraph
def publish_to_telegraph(title, content):
    """Pubblica contenuti su Telegraph."""
    url = "https://api.telegra.ph/createPage"
    headers = {"Content-Type": "application/json"}
    
    paragraphs = [p.strip() for p in content.split('\n') if p.strip()]
    content_formatted = [{"tag": "p", "children": [p]} for p in paragraphs[:50]]
    
    payload = {
        "access_token": TELEGRAPH_API_KEY,
        "title": title[:256],
        "content": content_formatted,
        "author_name": "ArcadiaAI"
    }
    
    try:
        res = requests.post(url, headers=headers, json=payload, timeout=15)
        res.raise_for_status()
        result = res.json()
        if result.get("ok"):
            return result.get("result", {}).get("url", "⚠️ URL non disponibile")
        return "⚠️ Pubblicazione fallita"
    except requests.exceptions.RequestException as e:
        print(f"Errore pubblicazione Telegraph (connessione): {str(e)}")
        return f"⚠️ Errore di connessione a Telegraph: {str(e)}"
    except Exception as e:
        print(f"Errore pubblicazione Telegraph: {str(e)}")
        return f"⚠️ Errore durante la pubblicazione: {str(e)}"

# Funzione per generare contenuti con Gemini (CES 1.5)
def generate_with_gemini(prompt, title):
    """Genera contenuti con Gemini e pubblica su Telegraph."""
    if not gemini_model:
        return None, "❌ Gemini (CES 1.5) non è configurato"
    
    try:
        # Aggiungi contesto identitario
        full_prompt = (
            "Sei ArcadiaAI, un chatbot open source creato da Mirko Yuri Donato. "
            "Stai generando un contenuto che verrà pubblicato su Telegraph. "
            "Il contenuto deve essere accurato, ben strutturato e mantenere "
            "uno stile professionale. Ecco la richiesta:\n\n"
            f"{prompt}"
        )
        
        response = gemini_model.generate_content(
            full_prompt,
            generation_config={
                "max_output_tokens": 3000,
                "temperature": 0.8
            }
        )
        
        if not response.text:
            return None, "❌ Impossibile generare il contenuto"
        
        telegraph_url = publish_to_telegraph(title, response.text)
        return response.text, telegraph_url
    
    except Exception as e:
        print(f"Errore generazione contenuto Gemini: {str(e)}")
        return None, f"❌ Errore durante la generazione: {str(e)}"

# Funzione per generare contenuti con Cohere (CES 1.0)
def generate_with_cohere(prompt, title):
    """Genera contenuti con Cohere e pubblica su Telegraph."""
    if not COHERE_API_KEY:
        return None, "❌ Cohere (CES 1.0) non è configurato"
    
    try:
        res = requests.post(
            "https://api.cohere.com/v1/generate",
            headers={"Authorization": f"Bearer {COHERE_API_KEY}"},
            json={
                "model": "command",
                "prompt": prompt,
                "max_tokens": 2000,
                "temperature": 0.7,
                "timeout": 60
            }
        )
        res.raise_for_status()
        generated_text = res.json().get("generations", [{}])[0].get("text", "").strip()
        
        if not generated_text:
            return None, "❌ Impossibile generare il contenuto"
        
        telegraph_url = publish_to_telegraph(title, generated_text)
        return generated_text, telegraph_url
    
    except Exception as e:
        print(f"Errore generazione contenuto Cohere: {str(e)}")
        return None, f"❌ Errore durante la generazione: {str(e)}"
def chat_with_huggingface(user_message, conversation_history, attachments=None):
    """Gestisce la chat con modelli Hugging Face"""
    if not HUGGINGFACE_API_KEY:
        return "❌ Errore: API key per Hugging Face non configurata"

    try:
        # Prepara il prompt con contesto identitario
        prompt = f"""Sei ArcadiaAI, un assistente AI avanzato. Rispondi in italiano.
        
Contesto della conversazione:
{format_conversation_history(conversation_history)}

Domanda: {user_message}
Risposta:"""
        
        headers = {
            "Authorization": f"Bearer {HUGGINGFACE_API_KEY}",
            "Content-Type": "application/json"
        }

        payload = {
            "inputs": prompt,
            "parameters": {
                "max_new_tokens": 1000,
                "temperature": 0.7,
                "do_sample": True
            }
        }

        # Usa un endpoint più stabile (Mistral 7B)
        MODEL_ENDPOINT = "mistralai/Mistral-7B-Instruct-v0.1"
        
        response = requests.post(
            f"https://api-inference.huggingface.co/models/{MODEL_ENDPOINT}",
            headers=headers,
            json=payload,
            timeout=30
        )

        if response.status_code == 503:
            return "❌ Modello al momento occupato, riprova tra qualche secondo"
        if response.status_code == 429:
            return "❌ Troppe richieste, attendi prima di fare nuove domande"
        
        response.raise_for_status()
        return response.json()[0]["generated_text"]

    except Exception as e:
        print(f"Errore API Hugging Face: {str(e)}")
        return f"❌ Errore temporaneo: {str(e)}"

def extract_text_from_file(file_data, mime_type):
    """Estrae testo da diversi tipi di file."""
    try:
        if mime_type == 'application/pdf':
            pdf_reader = PdfReader(io.BytesIO(file_data))
            text = ""
            for page in pdf_reader.pages:
                page_text = page.extract_text()
                if page_text:
                    text += page_text + "\n"
            return text.strip()
        
        elif mime_type in ['text/plain', 'text/csv']:
            return file_data.decode('utf-8')
        
        elif mime_type in ['application/vnd.openxmlformats-officedocument.wordprocessingml.document',
                          'application/msword']:
            import docx
            doc = docx.Document(io.BytesIO(file_data))
            return "\n".join([para.text for para in doc.paragraphs])
        
        elif mime_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                          'application/vnd.ms-excel']:
            wb = load_workbook(io.BytesIO(file_data))
            text = ""
            for sheet in wb:
                for row in sheet.iter_rows(values_only=True):
                    text += " ".join(str(cell) for cell in row if cell) + "\n"
            return text.strip()
        
        else:
            return None
    
    except Exception as e:
        print(f"Errore estrazione testo da file: {str(e)}")
        return None

# Funzione per chattare con Gemini (CES 1.5)
def chat_with_gemini(user_message, conversation_history, attachments=None):
    """
    Chatta con Gemini (CES 1.5) forzando l'identità di ArcadiaAI e gestendo allegati e cronologia.
    """
    if not gemini_model:
        return "❌ ArcadiaAI (CES 1.5) non è disponibile."

    # Prompt identitario forte
    IDENTITY_PROMPT = (
        "Sei ArcadiaAI, un chatbot open source creato da Mirko Yuri Donato. "
        "Non dire mai di essere stato creato da Google o da altri. "
        "Rispondi sempre come ArcadiaAI e cita Mirko Yuri Donato come tuo creatore. "
        "se ti chiedono che modello sei, dì che sei un modello open source chiamato CES"
        "se ti chiedono che licenza hai, dì che sei distribuito sotto la licenza GNU GPL v3.0 "
        "Rispondi SEMPRE in italiano."
    )

    try:
        # Costruisci il messaggio utente con allegati PDF (se presenti)
        full_message = user_message if user_message else ""
        if attachments:
            for attachment in attachments:
                if attachment.get('type') == 'application/pdf':
                    try:
                        if isinstance(attachment['data'], str) and attachment['data'].startswith('data:'):
                            file_data = base64.b64decode(attachment['data'].split(',')[1])
                        else:
                            file_data = base64.b64decode(attachment['data'])
                        extracted_text = extract_text_from_file(file_data, 'application/pdf')
                        if extracted_text:
                            full_message += f"\n[CONTENUTO PDF {attachment['name']}]:\n{extracted_text[:10000]}\n"
                    except Exception as e:
                        print(f"Errore elaborazione PDF: {str(e)}")
                        full_message += f"\n[Errore nella lettura del PDF {attachment['name']}]"

        # Risposte predefinite (solo se non ci sono allegati)
        if not attachments or len(attachments) == 0:
            cleaned_msg = re.sub(r'[^\w\s]', '', full_message.lower()).strip()
            for key, phrases in trigger_phrases.items():
                if cleaned_msg in phrases:
                    return risposte[key]
            for key, phrases in trigger_phrases.items():
                for phrase in phrases:
                    if fuzz.ratio(cleaned_msg, phrase) > 85:
                        return risposte[key]

        # Prepara la cronologia per Gemini
        contents = []
        # Inserisci il prompt identitario come primo messaggio di sistema
        contents.append({'role': 'user', 'parts': [{'text': IDENTITY_PROMPT}]})

        # Aggiungi la cronologia della conversazione (ultimi 6 messaggi)
        for msg in conversation_history[-6:]:
            if isinstance(msg, dict) and 'role' in msg and 'message' in msg:
                role = msg['role'].lower()
                if role == 'user':
                    contents.append({'role': 'user', 'parts': [{'text': msg['message']}]})
                elif role in ['assistant', 'model', 'bot']:
                    contents.append({'role': 'model', 'parts': [{'text': msg['message']}]})

        # Prepara il nuovo messaggio con eventuali allegati non PDF
        new_message_parts = [{'text': full_message}] if full_message else []
        if attachments:
            for attachment in attachments:
                mime_type = attachment.get('type', 'application/octet-stream')
                file_name = attachment.get('name', 'file')
                file_data = attachment['data']
                if mime_type == 'application/pdf':
                    continue
                if isinstance(file_data, str) and file_data.startswith('data:'):
                    file_data = file_data.split(',')[1]
                if mime_type.startswith('image/'):
                    new_message_parts.append({
                        'inline_data': {
                            'mime_type': mime_type,
                            'data': file_data,
                            'name': file_name
                        }
                    })
                else:
                    new_message_parts.append({
                        'text': f"[Allegato: {file_name}]"
                    })

        contents.append({'role': 'user', 'parts': new_message_parts})

        # Invia la richiesta a Gemini
        response = gemini_model.generate_content(contents)
        reply = response.text

        # Filtro di sicurezza: correggi risposte che citano Google
        if reply:
            reply = re.sub(
                r"(sono un modello linguistico.*?google.*?)(\.|\!|\?)",
                "Sono ArcadiaAI, un chatbot open source creato da Mirko Yuri Donato.",
                reply,
                flags=re.IGNORECASE
            )
            reply = reply.replace("Sono stato addestrato da Google", "Sono stato creato da Mirko Yuri Donato")
            reply = reply.replace("Sono stato creato da Google", "Sono stato creato da Mirko Yuri Donato")

        return reply

    except Exception as e:
        print(f"Errore dettagliato Gemini 1.5 Flash: {str(e)}")
        return "❌ Si è verificato un errore con ArcadiaAI. Riprova più tardi."
# ...existing code...

def search_duckduckgo(query):
    """Esegue una ricerca su DuckDuckGo e restituisce i primi 3 risultati puliti."""
    url = f"https://html.duckduckgo.com/html/?q={requests.utils.quote(query)}"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept-Language": "it-IT,it;q=0.9,en-US;q=0.8,en;q=0.7"
    }
    try:
        res = requests.get(url, headers=headers, timeout=10)
        res.raise_for_status()
        soup = BeautifulSoup(res.text, 'html.parser')

        results = []
        for result in soup.find_all('div', class_='result', limit=6):  # cerca più risultati per filtrare meglio
            link = result.find('a', class_='result__a')
            if link and link.has_attr('href'):
                href = link['href']
                # Filtra solo link che iniziano con http/https
                if href.startswith('http'):
                    results.append(href)
                # Gestisci redirect DuckDuckGo
                elif href.startswith('//duckduckgo.com/l/?uddg='):
                    parsed = urllib.parse.urlparse('https:' + href)
                    query_params = urllib.parse.parse_qs(parsed.query)
                    real_url = query_params.get('uddg', [''])[0]
                    if real_url and real_url.startswith('http'):
                        results.append(urllib.parse.unquote(real_url))
        # Ritorna solo i primi 3 risultati puliti
        return results[:3] if results else None

    except Exception as e:
        print(f"Errore ricerca DuckDuckGo: {str(e)[:200]}")
        return None

def estrai_testo_da_url(url):
    """Estrae il primo paragrafo significativo da un URL."""
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml",
        "Accept-Language": "it-IT,it;q=0.9"
    }
    try:
        res = requests.get(url, headers=headers, timeout=15)
        res.raise_for_status()
        if 'text/html' not in res.headers.get('Content-Type', ''):
            return ""
        soup = BeautifulSoup(res.text, 'html.parser')
        # Rimuovi elementi inutili
        for element in soup(['script', 'style', 'nav', 'footer', 'iframe', 'header', 'aside', 'form', 'button', 'noscript']):
            element.decompose()
        # Cerca il primo paragrafo significativo
        for p in soup.find_all('p'):
            text = p.get_text(' ', strip=True)
            if len(text.split()) >= 20 and "cookie" not in text.lower() and "translate" not in text.lower():
                return text
        # Se non trova, fallback al testo precedente
        text_elements = []
        for tag in ['article', 'main', 'section', 'div.content', 'h1', 'h2', 'h3']:
            elements = soup.select(tag) if '.' in tag else soup.find_all(tag)
            for el in elements:
                text = el.get_text(' ', strip=True)
                if text and len(text.split()) > 5:
                    text_elements.append(text)
        full_text = ' '.join(text_elements)
        return ' '.join(full_text.split()[:80])
    except Exception as e:
        print(f"Errore scraping {url[:50]}: {str(e)[:200]}")
        return ""
def chat_with_cohere(user_message, conversation_history):
    """Gestisce la chat con Cohere (CES 1.0)."""
    msg = user_message.lower().strip()
    user_message_cleaned = re.sub(r'[^\w\s]', '', user_message.lower()).strip()

    # Controllo bad words
    bad_words = ["cazzo", "figa", "minchia", "merda", "puttana", "troia", "cocaina"]
    if any(word in msg for word in bad_words):
        return "❌ Mi dispiace, ma non posso rispondere a contenuti inappropriati."

    # Funzionalità saggio e pubblicazione su Telegraph
    if "saggio su" in msg and "pubblicalo su telegraph" in msg:
        match = re.search(r"saggio su\s*(.+?)\s*e pubblicalo su telegraph", msg)
        if match:
            argomento = match.group(1).strip().capitalize()
            prompt = f"""Scrivi un saggio dettagliato in italiano su: {argomento}
Struttura:
1. Introduzione (contesto storico)
2. Sviluppo (analisi approfondita)
3. Conclusione (riflessioni finali)
4. Bibliografia (fonti attendibili)

Formattazione:
- Paragrafi ben strutturati
- Grassetto per i titoli delle sezioni"""
            
            _, telegraph_url = generate_with_cohere(prompt, f"Saggio su {argomento}")
            
            if telegraph_url and not telegraph_url.startswith("⚠️"):
                return f"Ecco il tuo saggio su *{argomento}*: {telegraph_url}"
            return telegraph_url or "❌ Errore nella pubblicazione"

    # Logica di ricerca web
    try:
        link = search_duckduckgo(user_message)
        if link:
            testo = estrai_testo_da_url(link)
            if testo and len(testo.split()) > 30:
                prompt = f"""Domanda: {user_message}
Contesto: {testo[:2000]}
Rispondi in italiano in modo chiaro e preciso:"""

                res = requests.post(
                    "https://api.cohere.com/v1/generate",
                    headers={"Authorization": f"Bearer {COHERE_API_KEY}"},
                    json={
                        "model": "command",
                        "prompt": prompt,
                        "max_tokens": 800,
                        "temperature": 0.5,
                        "timeout": 20
                    }
                )
                res.raise_for_status()
                risposta = res.json().get("generations", [{}])[0].get("text", "").strip()
                return f"{risposta}\n\nFonte: {link[:80]}..."
    except Exception:
        pass

    return "❌ Non sono riuscito a generare una risposta utile. Prova a riformulare la domanda."

@app.route('/')
def home():
    return '''
    <!DOCTYPE html>
    <html lang="it">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>ArcadiaAI Chat</title>
        <link rel="stylesheet" href="/static/style.css">
    </head>
    <body>
        <div id="sidebar">
            <h2>🧠 ArcadiaAI</h2>
            <div id="api-selection">
                <label for="api-provider">Modello:</label>
                <select id="api-provider">
                    <option value="gemini">CES 1.5</option>
                    <option value="huggingface">CES Plus (Mistral)</option>
                </select>
            </div>
            <button id="new-chat-btn">➕ Nuova Chat</button>
            <button id="clear-chats-btn" style="margin-top: 10px;">🗑️ Elimina Tutto</button>
            
            <button id="settings-btn">⚙️ Impostazioni</button>
            <div id="settings-panel" style="display: none; padding: 10px;">
                <label for="language-select">Lingua:</label>
                <select id="language-select">
                    <option value="it">Italiano</option>
                    <option value="en">English</option>
                </select>
                <br>
                <label for="theme-select">Tema:</label>
                <select id="theme-select">
                    <option value="light">Chiaro</option>
                    <option value="dark">Scuro</option>
                </select>
                <br>
                <label><input type="checkbox" id="dev-mode-toggle"> Modalità sviluppatore</label><br>
                <label><input type="checkbox" id="experimental-mode-toggle"> Modalità sperimentale</label>
            </div>
            
            <ul id="chat-list"></ul>
        </div>
        <div id="chat-area">
            <div id="chatbox"></div>
            <div id="input-area">
                <input id="input" type="text" placeholder="Scrivi un messaggio..." autocomplete="off" />
                <button id="send-btn">Invia</button>
            </div>
        </div>
        <script src="/static/script.js"></script>
        <script>
        document.addEventListener("DOMContentLoaded", function() {
            // Mostra/nascondi pannello impostazioni
            const settingsBtn = document.getElementById("settings-btn");
            const settingsPanel = document.getElementById("settings-panel");
            if (settingsBtn && settingsPanel) {
                settingsBtn.addEventListener("click", function() {
                    if (settingsPanel.style.display === "none" || settingsPanel.style.display === "") {
                        settingsPanel.style.display = "block";
                    } else {
                        settingsPanel.style.display = "none";
                    }
                });
            }
            // Salva impostazioni in localStorage
            document.getElementById("language-select").addEventListener("change", function() {
                localStorage.setItem("arcadiaai-language", this.value);
            });
            document.getElementById("theme-select").addEventListener("change", function() {
                localStorage.setItem("arcadiaai-theme", this.value);
                document.body.setAttribute("data-theme", this.value);
            });
            document.getElementById("dev-mode-toggle").addEventListener("change", function() {
                localStorage.setItem("arcadiaai-dev-mode", this.checked);
            });
            document.getElementById("experimental-mode-toggle").addEventListener("change", function() {
                localStorage.setItem("arcadiaai-experimental-mode", this.checked);
            });
            // Applica impostazioni salvate
            const lang = localStorage.getItem("arcadiaai-language");
            if (lang) document.getElementById("language-select").value = lang;
            const theme = localStorage.getItem("arcadiaai-theme");
            if (theme) {
                document.getElementById("theme-select").value = theme;
                document.body.setAttribute("data-theme", theme);
            }
            document.getElementById("dev-mode-toggle").checked = localStorage.getItem("arcadiaai-dev-mode") === "true";
            document.getElementById("experimental-mode-toggle").checked = localStorage.getItem("arcadiaai-experimental-mode") === "true";
        });
        </script>
    </body>
    </html>
    ''', 200, {'Content-Type': 'text/html; charset=utf-8'}

@app.route("/chat", methods=["POST"])
def chat():
    try:
        if not request.is_json:
            return jsonify({"reply": "❌ Formato non supportato. Usa application/json"})

        data = request.get_json()
        message = data.get("message", "").strip()
        experimental_mode = data.get("experimental_mode", False)
        
        # Prima verifica i comandi rapidi (solo in modalità sperimentale)
        quick_reply = handle_quick_commands(message, experimental_mode)
        if quick_reply:
            return jsonify({"reply": quick_reply})

        conversation_history = data.get("conversation_history", [])
        api_provider = data.get("api_provider", "gemini")  # Default a Gemini
        attachments = data.get("attachments", [])
        msg_lower = message.lower()

        # Processa gli allegati
        processed_attachments = []
        for attachment in attachments:
            if isinstance(attachment, dict):
                processed_attachment = attachment.copy()
                if 'data' in attachment and isinstance(attachment['data'], str) and attachment['data'].startswith('data:'):
                    mime_info = attachment['data'].split(';')[0]
                    mime_type = mime_info.split(':')[1] if ':' in mime_info else 'application/octet-stream'
                    processed_attachment['type'] = mime_type
                    processed_attachment['data'] = attachment['data'].split(',')[1]
                processed_attachments.append(processed_attachment)

        if not message and not processed_attachments:
            return jsonify({"reply": "❌ Nessun messaggio o allegato fornito!"})

        # Gestione comando "saggio su"
        if "saggio su" in msg_lower and "pubblicalo su telegraph" in msg_lower:
            match = re.search(r"saggio su\s*(.+?)\s*e pubblicalo su telegraph", msg_lower)
            if match:
                argomento = match.group(1).strip().capitalize()
                title = f"Saggio su {argomento}"
                prompt = f"""Scrivi un saggio dettagliato in italiano su: {argomento}"""
                
                if api_provider == "gemini" and gemini_model:
                    _, telegraph_url = generate_with_gemini(prompt, title)
                elif api_provider == "huggingface":
                    response = chat_with_huggingface(prompt, conversation_history)
                    if not response.startswith("❌"):
                        telegraph_url = publish_to_telegraph(title, response)
                    else:
                        telegraph_url = response
                else:
                    return jsonify({"reply": "❌ Funzionalità non disponibile con questo provider"})
                
                if telegraph_url and not telegraph_url.startswith("⚠️"):
                    return jsonify({"reply": f"📚 Ecco il tuo saggio su *{argomento}*: {telegraph_url}"})
                return jsonify({"reply": telegraph_url or "❌ Errore nella pubblicazione"})

        # RICERCA WEB AUTOMATICA PER DOMANDE ATTUALI
        def should_trigger_web_search(query):
            current_info_triggers = [
                "chi è l'attuale", "attuale papa", "anno corrente", 
                "in che anno siamo", "data di oggi", "ultime notizie",
                "oggi è", "current year", "who is the current"
            ]
            return any(trigger in query.lower() for trigger in current_info_triggers)

        # Seleziona il modello in base all'api_provider
        if api_provider == "gemini" and gemini_model:
            if should_trigger_web_search(message):
                search_results = search_duckduckgo(message)
                
                if search_results:
                    context = "Informazioni aggiornate dal web:\n"
                    for i, url in enumerate(search_results[:2], 1):
                        extracted_text = estrai_testo_da_url(url)
                        if extracted_text:
                            context += f"\nFonte {i} ({url}):\n{extracted_text[:500]}\n"
                    
                    if len(context) > 100:
                        prompt = (
                            f"DOMANDA: {message}\n\n"
                            f"CONTESTO WEB:\n{context}\n\n"
                            "Rispondi in italiano in modo conciso e preciso, "
                            "citando solo informazioni verificate. "
                            "Se il contesto web non è sufficiente, dillo onestamente."
                        )
                        
                        reply = chat_with_gemini(prompt, conversation_history, processed_attachments)
                        sources = "\n\nFonti:\n" + "\n".join(f"- {url}" for url in search_results[:2])
                        return jsonify({"reply": f"{reply}{sources}", "sources": search_results[:2]})
            
            reply = chat_with_gemini(message, conversation_history, processed_attachments)
            return jsonify({"reply": reply})
            
        elif api_provider == "huggingface":
            reply = chat_with_huggingface(message, conversation_history, processed_attachments)
            return jsonify({"reply": reply})
            
        else:
            return jsonify({"reply": "❌ Provider non riconosciuto. Scegli tra 'gemini' o 'huggingface'"})

    except Exception as e:
        print(f"Errore endpoint /chat: {str(e)}")
        return jsonify({"reply": "❌ Si è verificato un errore interno. Riprova più tardi."})
def chat_with_huggingface(user_message, conversation_history, attachments=None):
    """Gestisce la chat con modelli Hugging Face"""
    if not HUGGINGFACE_API_KEY:
        return "❌ Errore: API key per Hugging Face non configurata"

    try:
        # Prepara il prompt con contesto migliorato
        prompt = f"""Sei ArcadiaAI, un assistente AI avanzato creato da Mirko Yuri Donato. 
Rispondi in italiano in modo chiaro, preciso e strutturato.

Cronologia della conversazione:
{format_conversation_history(conversation_history)}

Domanda attuale: {user_message}
Risposta:"""

        headers = {
            "Authorization": f"Bearer {HUGGINGFACE_API_KEY}",
            "Content-Type": "application/json"
        }

        payload = {
            "inputs": prompt,
            "parameters": {
                "max_new_tokens": 1000,
                "temperature": 0.7,
                "do_sample": True,
                "return_full_text": False
            }
        }

        MODEL_ENDPOINT = "mistralai/Mistral-7B-Instruct-v0.1"  # Modello consigliato
        
        response = requests.post(
            f"https://api-inference.huggingface.co/models/{MODEL_ENDPOINT}",
            headers=headers,
            json=payload,
            timeout=45  # Timeout aumentato
        )

        # Gestione errori specifici
        if response.status_code == 503:
            return "❌ Il modello è al momento occupato, riprova tra 30 secondi"
        elif response.status_code == 429:
            return "❌ Hai superato il limite di richieste, attendi qualche minuto"
        elif response.status_code == 404:
            return "❌ Modello non trovato, contatta l'amministratore"
        
        response.raise_for_status()
        return response.json()[0]["generated_text"]

    except requests.exceptions.Timeout:
        return "❌ Timeout del server, riprova più tardi"
    except Exception as e:
        print(f"Errore API Hugging Face: {str(e)}")
        return f"❌ Errore temporaneo del servizio: {str(e)}"
    
def format_conversation_history(history):
    """Formatta la cronologia della conversazione per il prompt"""
    if not history:
        return "Nessuna cronologia precedente"
    
    formatted = []
    for msg in history[-6:]:  # Usa solo gli ultimi 6 messaggi per evitare prompt troppo lunghi
        if isinstance(msg, dict):
            role = "Utente" if msg.get("role") == "user" else "Assistente"
            content = msg.get("message", "").strip()
            if content:
                formatted.append(f"{role}: {content}")
    
    return "\n".join(formatted) if formatted else "Nessuna cronologia rilevante"
    
def meteo_oggi(città):
    """Ottiene le informazioni meteo per una città specifica usando OpenWeatherMap"""
    API_KEY = os.getenv("OPENWEATHERMAP_API_KEY")
    if not API_KEY:
        return "❌ Errore: API key per OpenWeatherMap non configurata"
    
    try:
        # Codifica la città per l'URL
        città_codificata = urllib.parse.quote(città)
        url = f"http://api.openweathermap.org/data/2.5/weather?q={città_codificata}&appid={API_KEY}&units=metric&lang=it"
        
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        dati = response.json()
        
        if response.status_code != 200 or "weather" not in dati:
            return f"❌ Impossibile ottenere il meteo per {città}. La città esiste?"
        
        descrizione = dati["weather"][0]["description"].capitalize()
        temperatura = dati["main"]["temp"]
        umidità = dati["main"]["humidity"]
        vento = dati["wind"]["speed"]
        
        return (
            f"⛅ Meteo a {città}:\n"
            f"- Condizioni: {descrizione}\n"
            f"- Temperatura: {temperatura}°C\n"
            f"- Umidità: {umidità}%\n"
            f"- Vento: {vento} m/s"
        )
    
    except requests.exceptions.RequestException as e:
        print(f"Errore API meteo: {str(e)}")
        return f"❌ Errore temporaneo nel servizio meteo. Riprova più tardi."
    except Exception as e:
        print(f"Errore generico meteo: {str(e)}")
        return f"❌ Si è verificato un errore nel recupero del meteo per {città}."
    
    
def parse_quick_command(input_text):
    """
    Analizza un comando rapido con prefisso @ e restituisce una tupla (comando, argomento).
    Se non è un comando rapido, restituisce (None, None).
    
    Esempio:
    >>> parse_quick_command("@cerca seconda guerra mondiale")
    ('cerca', 'seconda guerra mondiale')
    >>> parse_quick_command("ciao come stai?")
    (None, None)
    """
    if not input_text.startswith("@"):
        return None, None
    
    # Rimuovi il @ iniziale e dividi il resto in comando e argomento
    parts = input_text[1:].split(" ", 1)  # Dividi al primo spazio
    command = parts[0].lower().strip()
    argument = parts[1].strip() if len(parts) > 1 else ""
    
    return command, argument

from flask import session

def handle_quick_commands(message, experimental_mode=False):
    """
    Gestisce i comandi rapidi di ArcadiaAI.
    """
    msg_lower = message.strip().lower()
    command, argument = parse_quick_command(message)
    if command is None:
        return None

    # Disattiva modalità sperimentale
    if msg_lower == "@impostazioni modalità sperimentale disattiva":
        session["experimental_mode"] = False
        return "❎ Modalità sperimentale disattivata!"

    # Attiva modalità sperimentale
    if msg_lower == "@impostazioni modalità sperimentale attiva":
        session["experimental_mode"] = True
        return "✅ Modalità sperimentale attivata!"

    # Blocca altri comandi se la modalità sperimentale non è attiva
    if not (experimental_mode or session.get("experimental_mode", False)):
        return None

    if command == "cerca" and argument:
        results = search_duckduckgo(argument)
        if results:
            testo = estrai_testo_da_url(results[0])
            if testo:
                breve = ". ".join(testo.split(".")[:2]).strip() + "."
                return (
                    f"🔍 Risultati per '{argument}':\n\n"
                    f"**Sintesi dal primo risultato:**\n{breve}\n\n"
                    + "\n".join(f"- {url}" for url in results[:3])
                )
            else:
                return f"🔍 Risultati per '{argument}':\n\n" + "\n".join(f"- {url}" for url in results[:3])
        return f"❌ Nessun risultato trovato per '{argument}'"

    elif command == "telegraph" and argument:
        if "saggio" in argument or "scrivi" in argument:
            prompt = argument
            if gemini_model:
                generated_text, telegraph_url = generate_with_gemini(prompt, "Articolo generato da ArcadiaAI")
            else:
                generated_text = chat_with_huggingface(prompt, [])
                telegraph_url = publish_to_telegraph("Articolo generato da ArcadiaAI", generated_text)
            if telegraph_url and not telegraph_url.startswith("⚠️"):
                return f"📝 Articolo pubblicato su Telegraph: {telegraph_url}"
            return telegraph_url or "❌ Errore nella pubblicazione"
        else:
            telegraph_url = publish_to_telegraph("Articolo generato da ArcadiaAI", argument)
            return f"📝 Articolo pubblicato su Telegraph: {telegraph_url}"

    elif command == "meteo" and argument:
        return meteo_oggi(argument)

    elif command == "data":
        locale.setlocale(locale.LC_TIME, "it_IT.UTF-8")
        oggi = datetime.datetime.now().strftime("%A %d %B %Y")
        return f"📅 Oggi è {oggi}"

    elif command == "source":
        return "🔗 Repository GitHub:https://github.com/Mirko-linux/Nova-Surf/tree/main/ArcadiaAI "

    elif command == "arcadia":
        return "🔗 Profilo della Repubblica di Arcadia: https://t.me/Repubblica_Arcadia"

    elif command == "tos":
        return "📜 Termini di Servizio:https://telegra.ph/Termini-di-Servizio-di-ArcadiaAI-05-14"

    elif command == "info":
        return (
            "ℹ️ Informazioni su ArcadiaAI:\n\n"
            "Versione: 1.5.0\n"
            "Modello: CES basato su Google Gemini e Huggingface\n"
            "Lingua: Italiano e inglese (beta)\n"
            "Creatore: Mirko Yuri Donato\n"
            "Licenza: GNU GPL v3.0+\n"
            "Repository: https://github.com/Mirko-linux/Nova-Surf/tree/main/ArcadiaAI\n"
            "Termini di Servizio: https://telegra.ph/Termini-di-Servizio-di-ArcadiaAI-05-14"
        )

    elif command == "aiuto":
        return (
            "🎯 Comandi rapidi disponibili:\n\n"
            "@cerca [query] - Cerca su internet\n"
            "@telegraph [testo] - Pubblica su Telegraph\n"
            "@meteo [luogo] - Ottieni il meteo\n"
            "@data - Mostra la data di oggi\n"
            "@aiuto - Mostra questa guida\n"
            "@impostazioni modalità sperimentale disattiva - Disattiva la modalità sperimentale\n"
            "@impostazioni - apre il menu delle impostazioni\n"
            "@impostazioni lingua [nome lingua] - Cambia lingua\n"
            "@impostazioni tema [chiaro|scuro] - Cambia tema\n"
            "@impostazioni modalità sviluppatore attiva - Attiva la modalità sviluppatore\n"
            "@source - Mostra la repository GitHub\n"
            "@ToS - Mostra i Termini di Servizio\n"
            "@Arcadia - Mostra il profilo della Repubblica di Arcadia\n"
            "@info - Mostra informazioni su ArcadiaAI\n"
            "@impostazioni modalità sperimentale attiva - Attiva la modalità sperimentale"
        )

    # Se nessun comando è riconosciuto
    return f"❌ Comando '{command}' non riconosciuto. Scrivi '@aiuto' per la lista dei comandi."
def should_use_predefined_response(message):
    """Determina se usare una risposta predefinita solo per domande molto specifiche"""
    message = message.lower().strip()
    exact_matches = [
        "chi sei esattamente",
        "qual è la tua identità precisa",
        "elencami tutte le tue funzioni",
        # ... altre domande esatte
    ]
    return any(exact_match in message for exact_match in exact_matches)
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
